import sqlite3
from openpyxl import Workbook
import os
import datetime

# Obtener la ruta del directorio actual del script
current_folder = os.path.dirname(os.path.abspath(__file__))

# Nombre del archivo de la base de datos
db_filename = 'precios_competencia.db'
db_path = os.path.join(current_folder, db_filename)

# Verificar si la base de datos existe en el directorio actual
if not os.path.isfile(db_path):
    print("No se encontró la base de datos en el directorio actual.")
    exit()

# Crear la conexión con la base de datos SQLite
conn = sqlite3.connect(db_path)
c = conn.cursor()

# Obtener nombres de las tablas en la base de datos
c.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")
tablas = c.fetchall()

# Encontrar la fecha más reciente en todas las tablas
max_fecha = None
for tabla in tablas:
    nombre_tabla = tabla[0]
    c.execute(f"SELECT MAX(fecha) FROM {nombre_tabla}")
    fecha_maxima = c.fetchone()[0]
    if fecha_maxima:
        fecha_maxima = datetime.datetime.strptime(fecha_maxima, '%d/%m/%Y')
        if max_fecha is None or fecha_maxima > max_fecha:
            max_fecha = fecha_maxima

# Si no se encontró ninguna fecha, salir
if max_fecha is None:
    print("No se encontraron fechas en las tablas.")
    exit()

# Convertir la fecha más reciente al formato de cadena necesario
max_fecha_str = max_fecha.strftime('%d/%m/%Y')

# Crear un nuevo archivo Excel
wb = Workbook()

for tabla in tablas:
    nombre_tabla = tabla[0]
    ws = wb.create_sheet(title=nombre_tabla)  # Crear una nueva hoja con el nombre de la tabla
    
    # Filtrar por la fecha más reciente
    c.execute(f"""
        SELECT DISTINCT fecha, descripcion, precio 
        FROM {nombre_tabla} 
        WHERE fecha = ? 
        ORDER BY fecha DESC, descripcion ASC
    """, (max_fecha_str,))
    rows = c.fetchall()
    
    if rows:
        # Obtener los nombres de las columnas
        column_names = [description[0] for description in c.description]
        # Escribir los nombres de las columnas en la primera fila
        for col_index, col_name in enumerate(column_names):
            ws.cell(row=1, column=col_index+1, value=col_name)
        # Escribir los datos en las filas siguientes
        for row_index, row in enumerate(rows):
            for col_index, value in enumerate(row):
                ws.cell(row=row_index+2, column=col_index+1, value=value)  # Comenzar desde la segunda fila

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
    else:
        print(f"No se encontraron datos para la tabla {nombre_tabla} con fecha {max_fecha_str}.")

# Eliminar la hoja inicial por defecto "Sheet" solo si hay más de una hoja
if len(wb.sheetnames) > 1:
    wb.remove(wb["Sheet"])

# Guardar el archivo Excel en el mismo directorio que la base de datos
excel_file_path = os.path.join(current_folder, "precios_competencia.xlsx")
wb.save(excel_file_path)

# Cerrar la conexión con la base de datos
conn.close()


